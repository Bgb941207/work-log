import os
from openpyxl import load_workbook

# ======= 在這裡指定你的根資料夾路徑 =======
ROOT_FOLDER = r"C:\Users\USER\Desktop\ppt_2_excel_相對路徑\全興廠"
# =====================================

def add_hyperlinks_in_c(root_dir):
    """
    遞迴搜尋 root_dir 下所有 .xlsx/.xlsm 檔案，
    並把每個檔案中每張表格 C 欄的文字，設為超連結。
    """
    for dirpath, _, filenames in os.walk(root_dir):
        for fn in filenames:
            if fn.startswith("~$"):
                continue
            if not fn.lower().endswith((".xlsx", ".xlsm")):
                continue

            full_path = os.path.join(dirpath, fn)
            print(f"Processing: {full_path}")
            wb = load_workbook(full_path)
            changed = False

            for ws in wb.worksheets:
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=3)  # C 欄
                    addr = cell.value
                    if isinstance(addr, str) and addr.strip():
                        if getattr(cell, "hyperlink", None) == addr:
                            continue
                        cell.hyperlink = addr
                        cell.style = "Hyperlink"
                        changed = True

            if changed:
                wb.save(full_path)
                print("  → Saved with hyperlinks.")
            else:
                print("  → No changes needed.")
    print("Done.")

if __name__ == "__main__":
    add_hyperlinks_in_c(ROOT_FOLDER)
